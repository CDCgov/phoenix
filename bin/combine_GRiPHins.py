#!/usr/bin/env python3

#disable cache usage in the Python so __pycache__ isn't formed. If you don't do this using 'nextflow run cdcgov/phoenix...' a second time will causes and error
import sys
from packaging import version
sys.dont_write_bytecode = True
import glob
import os
import pandas as pd
import numpy as np
import argparse
import xlsxwriter as ws
from xlsxwriter.utility import xl_rowcol_to_cell
from openpyxl import load_workbook
from itertools import chain
from GRiPHin import order_ar_gene_columns, Combine_dfs, write_to_excel, convert_excel_to_tsv, print_df

##Makes a summary Excel file 
##Usage: >combine_GRiPHins.py -g1 141024_old_GRiPHin.xlsx -g2 141024_GRiPHin_Summary.xlsx --output Test --coverage 30 --parent_folder 141024 --bldb BLDB_20250327.csv --samplesheet samplesheet.valid.csv --remove_dups
## Written by Jill Hagey (qpk9@cdc.gov)

# Function to get the script version
def get_version():
    return "1.0.0"

def parseArgs(args=None):
    parser = argparse.ArgumentParser(description='''Script to create new griphin excel sheet by combining two griphin summaries. The -g2 is considered the "new" file and thus when 
    samples with data in the -g1 file will have values overwritten to be the values in the -g2 file. This script relies on functions in GRiPHin.py, so this script can only be used with GRiPHin.py from the same version of PHoeNIx.''')
    parser.add_argument('-g1', '--old_griphin', default=None, required=False, dest='griphin_old', help='The first griphin excel file to combine.')
    parser.add_argument('--remove_dups', default=False, action='store_true', required=False, dest='remove_dups', help='Pass this argument if there are duplicate sample IDs in your griphins and this will remove the old sample rows and replace them with the new data in the new griphin. Only works with -g1 -g2 and not --griphin_list.')
    parser.add_argument('-g2', '--new_griphin', default=None, required=False, dest='griphin_new', help='The second griphin excel file to combine.')
    parser.add_argument('--parent_folder', required=False, default=None, dest='parent_folder', help='Directory that will be used for filling in the column "parent_folder" excel sheet if needed for backward compatibility.')
    parser.add_argument('-o', '--output', required=False, default=None, dest='output', help='Name of output file default is GRiPHin_Summary.xlsx.')
    parser.add_argument('--scaffolds', dest="scaffolds", default=False, action='store_true', help='Turn on with --scaffolds to keep samples from failing/warnings/alerts that are based on trimmed data. Default is off.')
    parser.add_argument('-b', '--bldb', required=True, default=None, dest='bldb', help='Name of output file default is GRiPHin_Summary.xlsx.')
    parser.add_argument('-s', '--samplesheet', required=False, default=None, dest='samplesheet', help='samplesheet with sample,directory columns. Used to doublecheck sample names.')
    parser.add_argument('--griphin_list', required=False, default=None, nargs='?', const=True, type=str, dest='griphin_list', help='pass instead of -g1/-g2 when you want to combine more than 2 griphins. If you just pass --griphin_list the script assumes you have multiple griphin_summary.xlsx files in the current dir. You can also pass a csv that just has the full paths to the griphin files you want to combine.')
    parser.add_argument('--coverage', default=30, required=False, dest='set_coverage', help='The coverage cut off default is 30x.')
    parser.add_argument('--version', action='version', version=get_version())# Add an argument to display the version
    return parser.parse_args()

#set colors for warnings so they are seen
CRED = '\033[91m'+'\nWarning: '
CYELLOW = '\033[93m'
CEND = '\033[0m'

#pd.set_option('display.max_colwidth', 100)  # Set a reasonable limit instead of None
#pd.set_option('display.max_rows', None)
#pd.set_option('display.max_columns', None)

def df_has_index_dupes(df_to_check, label):
    # Check if the DataFrame has duplicate indices
    if df_to_check.index.duplicated().any():
        print(label, "has duplicate indices.")
    else:
        print(label, "does not have duplicate indices.")

def df_has_other_dupes(df_to_check, to_check, label):
    # Check if the DataFrame has duplicate indices
    if to_check in df_to_check.columns:
        if df_to_check[to_check].duplicated().any():
            print(label,'-',to_check, "has duplicate indices.")
        else:
            print(label,'-',to_check, "does not have duplicate indices.")
    else:
        print(label,'-',to_check, "column not found in DataFrame.")

def sort_samples_df(df):
    samples = df["WGS_ID"]
    try:
        # Try sorting by numeric part of the WGS_ID
        sorted_samples = sorted(samples, key=lambda x: int("".join([i for i in str(x) if i.isdigit()])))
    except ValueError:
        # Fallback to pure alphabetical sorting
        sorted_samples = sorted(samples)
    # Set index to WGS_ID temporarily and reorder
    df = df.set_index("WGS_ID")
    df = df.loc[sorted_samples]
    df.reset_index(inplace=True)  # keep WGS_ID as a column
    return df

def sort_columns_to_primary_ungrouped(df_to_sort):
    default_df = [ "UNI", "WGS_ID", "Parent_Folder", "Data_Location", "PHX_Version", "Minimum_QC_Check", "Minimum_QC_Issues", "Warnings", "Alerts", "Raw_Q30_R1_[%]", "Raw_Q30_R2_[%]", "Total_Raw_[reads]", "Paired_Trimmed_[reads]", "Total_Trimmed_[reads]", "Estimated_Trimmed_Coverage", "GC[%]", "Scaffolds", "Assembly_Length", "Assembly_Ratio", "Assembly_StDev", "Final_Taxa_ID", "Taxa_Source", "BUSCO_Lineage", "BUSCO_%Match", "Kraken_ID_Trimmed_Reads_%", "Kraken_ID_WtAssembly_%", "ShigaPass_Organism", "FastANI_Organism", "FastANI_%ID", "FastANI_%Coverage", "Species_Support_ANI", "Primary_MLST_Scheme", "Primary_MLST_Source", "Primary_MLST", "Primary_MLST_Alleles", "Secondary_MLST_Scheme", "Secondary_MLST_Source", "Secondary_MLST", "Secondary_MLST_Alleles", "MLST Clade", "Toxinotype", "Toxin-A_sub-type", "tcdA", "Toxin-B_sub-type", "tcdB", "tcdC_Variant", "tcdC other mutations", "tcdC", "tcdR", "tcdE", "PaLoc_NonTox_Variant", "PaLoc_NonTox other mutations", "PaLoc_NonTox", "cdtA", "cdtB", "cdtR_Variant", "cdtR other mutations", "cdtR", "cdtAB1", "cdtAB2", "gyrA known mutations", "gyrA other mutations", "gyrA", "gyrB known mutations", "gyrB other mutations", "gyrB", "dacS known mutations", "dacS other mutations", "dacS", "feoB known mutations", "feoB other mutations", "feoB", "fur known mutations", "fur other mutations", "fur", "gdpP known mutations", "gdpP other mutations", "gdpP", "glyC known mutations", "glyC other mutations", "glyC", "hemN known mutations", "hemN other mutations", "hemN", "hsmA known mutations", "hsmA other mutations", "hsmA","lscR known mutations", "lscR other mutations", "lscR", "marR known mutations", "marR other mutations", "marR", "murG known mutations", "murG other mutations", "murG", "nifJ known mutations", "nifJ other mutations", "nifJ", "PNimB known mutations", "PNimB other mutations", "PNimB", "rpoB known mutations", "rpoB other mutations", "rpoB", "rpoC known mutations", "rpoC other mutations", "rpoC", "sdaB known mutations", "sdaB other mutations", "sdaB", "thiH known mutations", "thiH other mutations", "thiH", "vanR known mutations", "vanR other mutations", "vanR", "vanS known mutations", "vanS other mutations", "vanS", "CEMB RT Crosswalk", "Inferred RT", "Probability", "ML Note", "Plasmid Info", "AR_Database", "No_AR_Genes_Found"] #, "HV_Database", "No_HVGs_Found", "Plasmid_Replicon_Database", "No_Plasmid_Markers" ]
    rest_cols = [col for col in df_to_sort.columns if col not in default_df]
    exist_cols = [col for col in default_df if col in df_to_sort.columns]
    sorted_df = df_to_sort[ exist_cols + rest_cols ]
    return sorted_df

def insert_shiga_column(df, add_col, insertion_col):
    if add_col not in df.columns and insertion_col in df.columns:
        df[add_col] = None # add column we need to the index. 
        cols = list(df.columns)
        # Remove to avoid duplication before reinserting
        cols.remove(add_col) #remove column name from the Python list of column names — it does not affect the actual DataFrame.
        insert_after = cols.index(insertion_col) + 1
        cols.insert(insert_after, add_col)
        df = df[cols]
    return df

##### May need to revisit this, VERY different from merge version
def add_and_concatenate(df1, df2, shiga):
    # since the first 2 dfs through this script might not have  - allowing backwards compatibility 
    if shiga:
        df1 = insert_shiga_column(df1, 'ShigaPass_Organism', "Kraken_ID_WtAssembly_%")
        df2 = insert_shiga_column(df2, 'ShigaPass_Organism', "Kraken_ID_WtAssembly_%")
    # allowing backwards compatibility for griphins that don't have PHX_Version column
    df1 = insert_shiga_column(df1, 'PHX_Version', "Data_Location")
    df2 = insert_shiga_column(df2, 'PHX_Version', "Data_Location")
    # Step 1: Preserve column order from df1 and append missing columns from df2
    all_columns = list(df1.columns) + [col for col in df2.columns if col not in df1.columns]
    # Step 1: Find the union of columns in both DataFrames
    #all_columns = set(df1.columns).union(set(df2.columns))
    # Step 2: Add missing columns to each DataFrame at once by reindexing
    df1 = df1.reindex(columns=all_columns)
    df2 = df2.reindex(columns=all_columns)
    # Step 3: Concatenate the two DataFrames row-wise
    result = pd.concat([df1, df2], ignore_index=True, sort=False)
    # Step 4: Reset the index and return the final DataFrame
    result.reset_index(drop=True, inplace=True)
    return result

def update_centar_column_names_lens(centar_df_column_names_final, centar_df_column_names):
    # Iterate over each corresponding pair of lists in centar_df_lens_final and centar_df_lens
    for i, (final_list, new_list) in enumerate(zip(centar_df_column_names_final, centar_df_column_names)):
        # Convert both lists to sets to get unique column names
        unique_columns = set(final_list).union(set(new_list))
        # Update centar_df_lens_final with the unique columns
        centar_df_column_names_final[i] = list(unique_columns)
    # After processing all lists, calculate the total number of unique names in each list
    total_unique_columns_lens = [len(columns) for columns in centar_df_column_names_final]
    return total_unique_columns_lens, centar_df_column_names_final

def read_excel(file_path, old_phoenix, reference_qc_df, reference_centar_df, samplesheet, previous_centar_df_lens, previous_centar_df_column_names, parent_folder=None):
    #get number of footer lines for each file to skip when reading in as a pd
    footer_lines = detect_footer_lines(file_path)
    # Read the Excel file, skipping the first row and using the second row as the header
    try: #check that this is an excel file
        df = pd.read_excel(file_path,
            skiprows=1,  # Skip the first header row
            header=0,    # Use the second row as the header
            skipfooter=footer_lines,engine='openpyxl')
        if parent_folder is not None:
            # Add a suffix number for each duplicate WGS_ID (starting from 1)
            df['UNI'] = df.groupby('WGS_ID').cumcount() + 1
        if 'Parent_Folder' not in df.columns:
            df = backwards_compatibility(df, parent_folder, file_path)
        # Use vectorized operations for speed to get UNI column
        df['UNI'] = df['Parent_Folder'] + '/' + df['Data_Location'] + '/' + df['WGS_ID']
        df = df[['UNI'] + [col for col in df.columns if col != 'UNI']]
    except Exception as e:
        raise ValueError(f"The input file is not a valid Excel file: {file_path}")
    # check that we have the files from the same entry # set that its a CDC PHOENIX run
    phoenix = "BUSCO_Lineage" not in df.columns
    if old_phoenix != phoenix: 
        raise ValueError(f"{CRED}The one griphin file in your set was produced from -entry CDC_PHOENIX and some were not wasn't. These files aren't compatible to combine. {CEND}")
    #check for centar - centar is true if "Toxin-A_sub-type" is found
    centar = "Toxin-A_sub-type" in df.columns or "Toxinotype" in df.columns 
    #check for shigapass - shiga is true if "ShigaPass_Organism" is found
    shiga = "ShigaPass_Organism" in df.columns
    #parse old griphin df
    #if samplesheet != None:
    #    samplesheet = pd.read_csv(samplesheet, header=0)
    #    if 'directory' in samplesheet.columns:
    #        samples = samplesheet['directory'].to_list()
    #    elif 'sample' in samplesheet.columns:
    #        samples =  samplesheet['sample'].tolist()
    #    #filters the DataFrame df_1 to include only the rows where the values in the 'WGS_ID' column are present in the samples list.
    #    df = df[df['UNI'].isin(samples)]
    # check that we have the files from the same entry
    if centar == True:
        ordered_centar_df, centar_df_lens, centar_df_column_names = split_centar_df(centar, file_path, set(df["UNI"]))
        df_qc, df_gene_hits_with_centar = split_dataframe(df,"Toxinotype")
        df_centar, df_gene_hits = split_dataframe(df_gene_hits_with_centar,"AR_Database")
        df_ar, df_pf_hv = split_dataframe(df_gene_hits,"HV_Database")
        df_hv, df_pf = split_dataframe(df_pf_hv,"Plasmid_Replicon_Database")
    else:
        centar_df_lens = previous_centar_df_lens # keep the numbers from the previous dataframe
        centar_df_column_names = previous_centar_df_column_names
        ordered_centar_df = pd.DataFrame(columns=["UNI"])
        df_qc, df_gene_hits = split_dataframe(df,"AR_Database")
        df_ar, df_pf_hv = split_dataframe(df_gene_hits,"HV_Database")
        df_hv, df_pf = split_dataframe(df_pf_hv,"Plasmid_Replicon_Database")
    # Ensure the first column is 'UNI'
    if (df_qc.columns[0] != 'UNI' and reference_qc_df.columns[0] != 'UNI'):
        raise ValueError("The first column in both dataframes must be 'UNI'")
    # Set WGS_ID as index
    # Set WGS_ID as index
    df1_qc = reference_qc_df.set_index('UNI')
    ref_qc = df_qc.set_index('UNI')
    # Identify samples to add and print them
    samples_to_add = ref_qc.index.difference(df1_qc.index)
    #make sure the order of the ar genes is correct
    order_ar_df = order_ar_gene_columns(df_ar, True)
    return df_qc, order_ar_df, df_pf, df_hv, phoenix, shiga, centar, ordered_centar_df, centar_df_lens, centar_df_column_names

def combine_centar(ordered_centar_df_1, centar_df_lens_1, centar_df_column_names_1, ordered_centar_df_2, centar_df_lens_2, centar_df_column_names_2):
    # Combine the ordered_centar_df DataFrames
    reordered_centar_df = pd.concat([ordered_centar_df_1, ordered_centar_df_2], ignore_index=False, sort=False, copy=False)
    # Combine the centar_df_lens lists
    centar_df_lens = [max(lens) for lens in zip(centar_df_lens_1, centar_df_lens_2)]
    #separate dataframes
    RB_type = [ "CEMB RT Crosswalk", "Inferred RT", "Probability", "ML Note", "Plasmid Info" ]
    RB_type_col = [col for col in reordered_centar_df.columns if any(substring in col for substring in RB_type) ]
    A_B_Tox = [ "Toxinotype", "Toxin-A_sub-type", "tcdA", "Toxin-B_sub-type", "tcdB"]
    A_B_Tox_col = [col for col in reordered_centar_df.columns if any(substring in col for substring in A_B_Tox) ]
    other_Tox = [ "tcdC_Variant", "tcdC other mutations", "tcdC", "tcdR", "tcdE", "cdtA", 'cdtR_Variant', "cdtR other mutations", "cdtB", "cdtR", "cdtAB1", "cdtAB2", "PaLoc_NonTox_Variant", "PaLoc_NonTox other mutations","PaLoc_NonTox" ]
    other_Tox_col = [col for col in reordered_centar_df.columns if any(substring in col for substring in other_Tox) ]
    mutants = [ 'gyrA','gyrB','dacS','feoB','fur','gdpP','glyC','hemN','hsmA','lscR','marR', 'murG','nifJ','PNimB','rpoB', 'rpoC','sdaB','thiH', 'vanR', 'vanS' ]
    dirty_mutations_col = [col for col in reordered_centar_df.columns if any(substring in col for substring in mutants) ]
    mutations_col = [ col for col in dirty_mutations_col if '(' not in col ]
    # List of mutation names to remove
    #mutations_to_remove = ['tcdC other mutations', 'cdtR other mutations', 'PaLoc_NonTox other mutations']
    # Remove each mutation name if it exists in mutations_col
    #mutations_col = [mutation for mutation in mutations_col if mutation not in mutations_to_remove]
    existing_columns_in_order = ['UNI'] + A_B_Tox_col + other_Tox_col + mutations_col + RB_type_col
    ordered_centar_df = reordered_centar_df[existing_columns_in_order]
    centar_column_packages = []
    centar_column_packages.append(A_B_Tox_col)
    centar_column_packages.append(other_Tox_col)
    centar_column_packages.append(mutations_col)
    centar_column_packages.append(RB_type_col)
    # Combine the centar_df_column_names lists
    return ordered_centar_df, centar_df_lens, centar_column_packages



def combine_gene_dataframes(old_df, new_df):
    # Ensure the first column is 'WGS_ID'
    #if new_df.columns[0] != 'WGS_ID' or old_df.columns[0] != 'WGS_ID':
    if new_df.columns[0] != 'UNI' and old_df.columns[0] != 'UNI':
        raise ValueError("The first column in both dataframes must be either 'UNI'")
    df1_gene = old_df.set_index('UNI')
    df2_gene = new_df.set_index('UNI')
    # Identify samples to add and print them
    samples_to_add = df2_gene.index.difference(df1_gene.index)
    # Combine dataframes, prioritizing new_df values and aligning columns
    combined_df = df1_gene.combine_first(df2_gene)
    # Add missing columns from new_df to old_df
    columns_to_add = [col for col in df2_gene.columns if col not in df1_gene.columns and col not in ["AR_Database", "HV_Database","Plasmid_Replicon_Database", "No_AR_Genes_Found", "No_HVGs_Found", "No_Plasmid_Markers", "WGS_ID", "UNI"]]
    for col in columns_to_add:
        combined_df[col] = combined_df[col].fillna(df2_gene[col])
    # Revert index to column
    combined_df.reset_index(inplace=True)
    return combined_df, samples_to_add

def combine_qc_dataframes(df1_qc, df2_qc):
    # Convert the 'WGS_ID' columns to the index to facilitate updating
    #df_has_other_dupes(df1_qc, 'WGS_ID', 'df1_qc')
    #df_has_other_dupes(df2_qc, 'WGS_ID', 'df2_qc')
    df1 = df1_qc.set_index('UNI', inplace=False)
    df2 = df2_qc.set_index('UNI', inplace=False)
    # Update df1_qc with matching rows from df2_qc
    df1.update(df2)
    # Append non-matching rows from df2_qc to df1_qc using pd.concat
    combined_df = pd.concat([df1, df2[~df2.index.isin(df1.index)]], copy=False)
    
    #df_has_other_dupes(combined_df, 'WGS_ID', 'combined df IN CQD')
    # Reset the index to restore the 'UNI' column
    combined_df.reset_index(inplace=True)
    #df_has_other_dupes(combined_df, 'WGS_ID', 'combined_df after index reset')
    combined_ordered_df = sort_columns_to_primary_ungrouped(combined_df)

    # Overlord suggestoins on trying to filter duplicate rowws by WGS_ID
    # Combine the DataFrames
    #df_combined = pd.concat([df1, df2], ignore_index=True)

    # Ensure 'PHX_Version' exists
    if 'PHX_Version' not in combined_ordered_df.columns:
        combined_ordered_df['PHX_Version'] = None

    # Normalize PHX_Version values
    def clean_version(ver):
        if pd.isna(ver):
            return version.parse("0.0.0")
        ver = ver.replace("-dev", ".dev0") # or ".dev1" if you want it later than dev0
        if ver[0:2] == "v.":
            ver=ver[2:]
        elif ver[0] == "v":
            ver=ver[1:]
        return version.parse(str(ver).strip())

    # Track removed rows
    removed_unis = []

    def resolve_group(group):
        #print(f"\nResolving group for WGS_ID: {group['WGS_ID'].iloc[0]}")
        #print(f"Group shape: {group.shape}")
        check_cols = [ 'Raw_Q30_R1_[%]', 'Raw_Q30_R2_[%]', 'Total_Raw_[reads]', 'Paired_Trimmed_[reads]', 'Total_Trimmed_[reads]', 'Estimated_Trimmed_Coverage', 'GC[%]']
        #print(f"QC uniqueness per column:\n{group[check_cols].nunique()}")
        if group[check_cols].nunique().le(1).all():
            group = group.copy()
            group['PHX_Version_Clean'] = group['PHX_Version'].apply(clean_version)
            sorted_group = group.sort_values(by='PHX_Version_Clean', ascending=False)
            kept = sorted_group.head(1)
            removed = sorted_group.iloc[1:]
            removed_unis.extend(removed['UNI'].tolist())
            return kept
        else:
            return group

    deduped_df = (
        combined_ordered_df.groupby('WGS_ID', group_keys=False)
                .apply(resolve_group)
                .drop(columns='PHX_Version_Clean', errors='ignore')
                .reset_index(drop=True)
    )

    # Logging as before
    if removed_unis:
        print("Removed duplicate rows with the following UNI values (due to matching QC fields and lower PHX_Version):")
        for u in removed_unis:
            print(" -", u)
    else:
        print("No duplicate rows removed.")

    return deduped_df#combined_ordered_df

def make_empty_species_specific_df(species_specific_df, other_df, start_col, end_col):
    """Building a empty pandas dataframe with the right column names to be able to combine correctly"""
    #first we create and empty dataframe with the columns from species_specific_df and the WGS_ID and UNI from other_df
    # Extract row names from other_df
    row_names = other_df.index
    # Find the indices of the start and end columns
    start_idx = species_specific_df.columns.get_loc(start_col)
    end_idx = species_specific_df.columns.get_loc(end_col) + 1  # +1 to make it inclusive
    selected_columns = list(species_specific_df.columns[start_idx:end_idx])
    # Step 2: Columns from other_df - need to keep the id information from the first df
    other_df_columns = [ "UNI", "WGS_ID" ]
    # Step 3: Initialize the new DataFrame with WGS_ID and UNI data from other_df
    new_df = other_df[other_df_columns].copy()
    # Step 4: Add the remaining columns from species_specific_df, filled with empty strings
    for col in selected_columns:
        new_df[col] = ''
    # now we need to create the list of lengths for each set of species specific information
    #define centar headlines
    centar_headlines=[['Toxinotype', 'Toxin-A_sub-type', 'tcdA', 'Toxin-B_sub-type', 'tcdB'], ['tcdC_Variant', 'tcdC other mutations', 'tcdC', 'tcdR', 'tcdE', 'cdtA', 'cdtB', 'cdtR_Variant', 'cdtR other mutations', 'cdtR', 'cdtAB1', 'cdtAB2', 'PaLoc_NonTox_Variant', 'PaLoc_NonTox other mutations', 'PaLoc_NonTox'], ['gyrA known mutations', 'gyrA other mutations', 'gyrA', 'gyrB known mutations', 'gyrB other mutations', 'gyrB', 'dacS known mutations', 'dacS other mutations', 'dacS', 'feoB known mutations', 'feoB other mutations', 'feoB', 'fur known mutations', 'fur other mutations', 'fur', 'gdpP known mutations', 'gdpP other mutations', 'gdpP', 'glyC known mutations', 'glyC other mutations', 'glyC', 'hemN known mutations', 'hemN other mutations', 'hemN', 'hsmA known mutations', 'hsmA other mutations', 'hsmA', 'lscR known mutations', 'lscR other mutations', 'lscR', 'marR known mutations', 'marR other mutations', 'marR', 'murG known mutations', 'murG other mutations', 'murG', 'nifJ known mutations', 'nifJ other mutations', 'nifJ', 'PNimB known mutations', 'PNimB other mutations', 'PNimB', 'PNimB |','rpoB known mutations', 'rpoB other mutations', 'rpoB', 'rpoC known mutations', 'rpoC other mutations', 'rpoC', 'sdaB known mutations', 'sdaB other mutations', 'sdaB', 'thiH known mutations', 'thiH other mutations', 'thiH', 'vanR known mutations', 'vanR other mutations', 'vanR', 'vanS known mutations', 'vanS other mutations', 'vanS'], ['CEMB RT Crosswalk', 'Inferred RT', 'Probability', 'ML Note', 'Plasmid Info']]
    new_df_lens=[0,0,0,0]
    found_headlines = []
    found_headlines_list = [[],[],[],[]]
    #loop through centar_headlines sections
    for section in range(0,len(centar_headlines)):
        #loop through centar column names in the section
        for liner in centar_headlines[section]:
            # check if the column name from the list is found in the new_df
            if liner in new_df.columns.tolist():
                found_headlines.append(liner)
                found_headlines_list[section].append(liner)
                if liner != 'WGS_ID' and liner != 'UNI':
                    new_df_lens[section]+=1 # don't count this so the #s are correct for griphin merging
    return new_df, new_df_lens, selected_columns.remove("MLST Clade")

def split_dataframe(df, split_column):
    # Ensure the first column is 'UNI'
    if df.columns[0] != 'UNI':
        raise ValueError("The first column must be 'UNI'")
    # Find the index of the split_column
    split_index = df.columns.get_loc(split_column)
    # Create two DataFrames based on the split index, including 'UNI' at the beginning FOR BOTH
    df_before = df.iloc[:, [0] + list(range(1, split_index))]  # All columns before the split column + 'WGS_ID'
    #df_before = df.iloc[:, list(range(1, split_index))]  # All columns before the split column + 'WGS_ID'
    df_after = df.iloc[:, [0] + list(range(split_index, df.shape[1]))]  # The split column and all columns after it + 'WGS_ID'
    return df_before, df_after

def split_centar_df(centar, excel, sample_names):
    footer_lines1 = detect_footer_lines(excel)
    #first open the excel sheet using the merged headers, this will give us the columns that all go with that header in the centar_df_lens and centar_df_column_names below
    df = pd.read_excel(excel,
        header=[0, 1],    # Use the 2nd row as the header
        skipfooter=footer_lines1,engine='openpyxl')
    if centar == True:
        df_1 = pd.read_excel(excel,
            skiprows=1,  # Skip the first header row
            header=0,    # Use the second row as the header
            skipfooter=footer_lines1 ,engine='openpyxl', dtype={'WGS_ID': str,'Parent_Folder': str,'Data_Location': str}) # Specifying dtypes prevents pandas from using more memory than necessary for each column.
        # Use vectorized operations for speed to creating UNI column for speed and lower memory usage
        df_1['UNI'] = df_1['Parent_Folder'] + '/' + df_1['Data_Location'] + '/' + df_1['WGS_ID']
        df_1 = df_1[['UNI'] + [col for col in df_1.columns if col != 'UNI']]
        # remove duplicate samplesheet
        df_1 = df_1[df_1['UNI'].isin(sample_names)]
        # Get the column names
        columns = df_1.columns.tolist()
        # Find the indices of the specified columns
        ## Might need to fiddle with MLST Clade here eventually...might not if original griphins get made right
        start_index = columns.index('Toxinotype')
        end_index = columns.index('AR_Database')
        # Get the columns between the specified columns
        columns_between = [columns[0]] + columns[start_index:end_index]
        # Subset the DataFrame
        ordered_centar_df = df_1[columns_between]
        ordered_centar_df = sort_columns_to_primary_ungrouped(ordered_centar_df)
        #get length of sub dataframes - using the header to determine how many columns are in each section
        centar_df_lens = [ len(df['Toxin A/B Variants'].columns), len(df['Other Toxins'].columns), len(df['C. difficile Specific AR Mutations'].columns), len(df['ML Predicted Ribotype'].columns) ] # have to have it in a list for sum() later
        centar_df_column_names = [ df['Toxin A/B Variants'].columns, df['Other Toxins'].columns, df['C. difficile Specific AR Mutations'].columns, df['ML Predicted Ribotype'].columns ] # have to have it in a list for sum() later
    else:
        centar_df_lens = [0,0,0,0] # we sum later so needs to be a list of numbers
        centar_df_column_names = [[],[],[],[]]
        ordered_centar_df = pd.DataFrame(columns=["UNI"])
    return ordered_centar_df, centar_df_lens, centar_df_column_names

def check_column_presence(df1, df1_path, df2, df2_path):
    df1_has_column = "BUSCO_Lineage" not in df1.columns
    df2_has_column = "BUSCO_Lineage" not in df2.columns
    if df1_has_column == df2_has_column:
        phoenix = df1_has_column
    elif df1_has_column == True and df2_has_column == False:
        phoenix = False
        raise ValueError(f"{CRED}The new griphin file was produced from -entry CDC_PHOENIX and the old griphin summary wasn't. These files aren't compatible.{CEND}")
    else:
        raise ValueError(f"{CRED}The old griphin file was produced from -entry CDC_PHOENIX and the new griphin summary wasn't. These files aren't compatible.{CEND}")
    #check for centar
    if "Toxinotype" in df1.columns or "MLST Clade" in df1.columns or "Toxinotype" in df2.columns or "MLST Clade" in df2.columns:
        if "Toxinotype" not in df1.columns or "MLST Clade" not in df1.columns:
            centar_1 = False
        else:
            centar_1 = True
        if "Toxinotype" not in df2.columns or "MLST Clade" not in df2.columns:
            centar_2 = False
        else:
            centar_2 = True
    else:
        centar_1 = centar_2 = False
    if centar_1 == True or centar_2 == True:
        all_centar = True
    else:
        all_centar = False
    #check for shigapass
    if "ShigaPass_Organism" in df1.columns or "ShigaPass_Organism" in df2.columns:
        shiga = True
    else:
        shiga = False
    return phoenix, shiga, centar_1, centar_2, all_centar

def remove_dup_rows(old_griphin_df, new_griphin_df):
    #omg why is scicomp like this....
    #old_griphin_df["UNI"] = old_griphin_df["UNI"].str.replace("/scicomp/groups/", "/scicomp/groups-pure/", regex=False)
    #new_griphin_df["UNI"] = new_griphin_df["UNI"].str.replace("/scicomp/groups/", "/scicomp/groups-pure/", regex=False)
    #remove dups
    old_griphin_df = old_griphin_df[~old_griphin_df['UNI'].isin(new_griphin_df['UNI'])]
    return old_griphin_df

def backwards_compatibility(df, parent_folder, file_path):
    #check if parent_folder is present
    #first print warning letting the user know. 
    print(CRED + f"{file_path} appears to be from an older version of PHoeNIx please use --parent_folder and pass the location of where the PHoeNIx output is found." + CEND)
    if parent_folder is not None:
        df['Parent_Folder'] = parent_folder
    else:
        df['Parent_Folder'] = ''
    if "Data_Location" not in df.columns:
        df['Data_Location'] = os.path.basename(os.path.dirname(parent_folder))
    return df

def detect_footer_lines(file_path, sheet_name=0):
    # Load the workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name if isinstance(sheet_name, str) else wb.sheetnames[sheet_name]]
    footer_lines = 3 #set to 3 because review by row has data in another cell and there are two all blank spacing rows
    is_footer_section = False
    # Count footer lines
    for row in reversed(tuple(ws.rows)):
        first_cell = row[0].value
        other_cells = [cell.value for cell in row[1:]]
        # Check if the first cell has text and all other cells are empty
        if isinstance(first_cell, str) and all(cell is None for cell in other_cells):
            footer_lines += 1
            is_footer_section = True
        elif is_footer_section:
            break
    return footer_lines

def read_excels(file_path1, file_path2, samplesheet, remove_dups, parent_folder):
    #get number of footer lines for each file to skip when reading in as a pd
    footer_lines1 = detect_footer_lines(file_path1)
    footer_lines2 = detect_footer_lines(file_path2)
    # Read the Excel file, skipping the first row and using the second row as the header
    try: #check that this is an excel file
        df_1 = pd.read_excel(file_path1,
            skiprows=1,  # Skip the first header row
            header=0,    # Use the second row as the header
            skipfooter=footer_lines1,engine='openpyxl')
        if 'Parent_Folder' not in df_1.columns:
            df_1 = backwards_compatibility(df_1, parent_folder, file_path1)
        #rename columns for backwards compatibility
        df_1 = df_1.rename(columns={'Kraken_ID_Raw_Reads_%': 'Kraken_ID_Trimmed_Reads_%'})
        # Use vectorized operations for speed to creating UNI column
        # Pre-compute the parent folder replacement
        df_1['Parent_Folder'] = df_1['Parent_Folder'].str.replace("/scicomp/groups/", "/scicomp/groups-pure/", regex=False)
        # Then create UNI using string formatting which is more memory efficient
        df_1['UNI'] = df_1['Parent_Folder'] + '/' + df_1['Data_Location'] + '/' + df_1['WGS_ID']
        df_1 = df_1[['UNI'] + [col for col in df_1.columns if col != 'UNI']]
    except Exception as e:
        raise ValueError(f"The input file is not a valid Excel file: {file_path1}")
    try: #check that this is an excel file
        df_2 = pd.read_excel(file_path2,
            skiprows=1,  # Skip the first header row
            header=0,    # Use the second row as the header
            skipfooter=footer_lines2,engine='openpyxl')
            #check if parent folder is present
        if 'Parent_Folder' not in df_2.columns:
            df_2 = backwards_compatibility(df_2, parent_folder, file_path2)
        #rename columns for backwards compatibility
        df_2 = df_2.rename(columns={'Kraken_ID_Raw_Reads_%': 'Kraken_ID_Trimmed_Reads_%'})
        # Pre-compute the parent folder replacement
        df_2['Parent_Folder'] = df_2['Parent_Folder'].str.replace("/scicomp/groups/", "/scicomp/groups-pure/", regex=False)
        # Then create UNI using string formatting which is more memory efficient and Use vectorized operations for speed to creating UNI column
        df_2['UNI'] = df_2['Parent_Folder'] + '/' + df_2['Data_Location'] + '/' + df_2['WGS_ID']
        df_2 = df_2[['UNI'] + [col for col in df_2.columns if col != 'UNI']]
    except Exception as e:
        raise ValueError(f"The input file is not a valid Excel file: {file_path2}")
    #drop duplicate samples from the old dataframe
    if remove_dups == True:
        #remove duplicate rows -- in the larger phx pipeline the things that are duplicates will be updated
        df_1 = remove_dup_rows(df_1, df_2)
    # because the dir is pulled into the module when making a griphin file some samples that aren't in the samplesheet can end up in the griphin file so we will remove those
    #if samplesheet != None:
    #    samplesheet = pd.read_csv(samplesheet, header=0)
    #    if 'directory' in samplesheet.columns:
    #        samples = samplesheet['directory'].to_list()
    #    elif 'sample' in samplesheet.columns:
    #        samples =  samplesheet['sample'].tolist()
    #    df_1 = df_1[df_1['UNI'].isin(samples)]
    #    df_2 = df_2[df_2['UNI'].isin(samples)]
    # check that we have the files from the same entry and if species specific columns are there
    phoenix, shiga, centar_1, centar_2, all_centar = check_column_presence(df_1, file_path1, df_2, file_path2)
    # We next start to split up the dataframes to go into already created functions in GRiPHin.py
    if centar_1 == True: #if first dataframe has Centar information then split that df as well as the others
        ordered_centar_df_1, centar_df_lens_1, centar_df_column_names_1 = split_centar_df(centar_1, file_path1, set(df_1["UNI"]))
        df1_qc, df1_gene_hits_with_centar = split_dataframe(df_1,"Toxinotype")
        df1_centar, df1_gene_hits = split_dataframe(df1_gene_hits_with_centar,"AR_Database")
        df1_ar, df1_pf_hv = split_dataframe(df1_gene_hits,"HV_Database")
        df1_hv, df1_pf = split_dataframe(df1_pf_hv,"Plasmid_Replicon_Database")
    else:
        #check if df_2 also has centar information. if so add blank centar info for the columns there to be able to combine, otherwise we will return an empty dataframe
        if centar_2 == True:
            #make an empty dataframe with all the species specific columns in centar_2 so we can combine them
            ordered_centar_df_1, centar_df_lens_1, centar_df_column_names_1 = make_empty_species_specific_df(df_2, df_1, "MLST Clade", "ML Note")
        else:
            #make an empty dataframe no columns we just have to have it be zeros to pass to the write_excel() function
            ordered_centar_df_1, centar_df_lens_1, centar_df_column_names_1 = split_centar_df(centar_1, file_path1, set(df_1["UNI"]))
        df1_qc, df1_gene_hits = split_dataframe(df_1,"AR_Database")
        df1_ar, df1_pf_hv = split_dataframe(df1_gene_hits,"HV_Database")
        df1_hv, df1_pf = split_dataframe(df1_pf_hv,"Plasmid_Replicon_Database")
    #doing the same splitting for the new dataframe
    if centar_2 == True:
        ordered_centar_df_2, centar_df_lens_2, centar_df_column_names_2 = split_centar_df(centar_2, file_path2, set(df_2["UNI"]))
        df2_qc, df2_gene_hits_with_centar = split_dataframe(df_2,"Toxinotype")
        df2_centar, df2_gene_hits = split_dataframe(df2_gene_hits_with_centar,"AR_Database")
        df2_ar, df2_pf_hv = split_dataframe(df2_gene_hits,"HV_Database")
        df2_hv, df2_pf = split_dataframe(df2_pf_hv,"Plasmid_Replicon_Database")
    else:
        #check if df_2 also has centar information. if so add blank centar info for the columns there to be able to combine
        #ordered_centar_df_1, centar_df_lens_1, centar_df_column_names_1 = split_centar_df(centar_1, file_path1)
        if centar_1 == True:
            #make an empty dataframe with all the species specific columns in centar_1 so we can combine them
            ordered_centar_df_2, centar_df_lens_2, centar_df_column_names_2 = make_empty_species_specific_df(df_1, df_2, "MLST Clade", "ML Note")
        else:
            #make an empty dataframe no columns we just have to have it be zeros to pass to the write_excel() function
            ordered_centar_df_2, centar_df_lens_2, centar_df_column_names_2 = split_centar_df(centar_2, file_path2, set(df_2["UNI"]))
        df2_qc, df2_gene_hits = split_dataframe(df_2,"AR_Database")
        df2_ar, df2_pf_hv = split_dataframe(df2_gene_hits,"HV_Database")
        df2_hv, df2_pf = split_dataframe(df2_pf_hv,"Plasmid_Replicon_Database")
    #combine qc columns
    combined_df_qc = combine_qc_dataframes(df1_qc, df2_qc)
    #combine ar columns
    combined_df_ar, samples_to_add = combine_gene_dataframes(df1_ar, df2_ar)
    #make sure the order of the ar genes is correct
    order_combined_ar_df = order_ar_gene_columns(combined_df_ar, True)
    print(CYELLOW + "\nAdding sample(s) to the GRiPHin summary:", samples_to_add.tolist(),"\n" + CEND)
    # combine pf and hv dataframes
    combined_df_pf, samples_to_add = combine_gene_dataframes(df1_pf, df2_pf)
    combined_df_hv, samples_to_add = combine_gene_dataframes(df1_hv, df2_hv)
    ordered_centar_df, centar_df_lens, centar_df_column_names = combine_centar(ordered_centar_df_1, centar_df_lens_1, centar_df_column_names_1, ordered_centar_df_2, centar_df_lens_2, centar_df_column_names_2)
    return combined_df_qc, order_combined_ar_df, combined_df_pf, combined_df_hv, phoenix, shiga, all_centar, ordered_centar_df, centar_df_lens, centar_df_column_names

def main():
    # looping through excel files
    args = parseArgs()
    #figure out the name of the output file 
    if args.output != None:
        output_file = args.output
    else:
        # Derive output file name from input file name
        output_file = args.griphin_new.replace("_GRiPHin_Summary.xlsx", "")
    # checking what the input type is
    if args.griphin_old != None and args.griphin_new != None: # only two files being combined
        combined_df_qc_final, combined_df_ar_final, combined_df_pf_final, combined_df_hv_final, phoenix_final, shiga_final, centar_final, ordered_centar_df_final, centar_df_lens_final, centar_df_column_names_final = read_excels(args.griphin_old, args.griphin_new, args.samplesheet, args.remove_dups, args.parent_folder)
    else:
        # Will find if griphin_list has an argument with it or not. Will build a griphin list if one is not given
        if args.griphin_list == True:
            # old_GRiPHin extension is for species specific pipeline(s)
            griphin_files = [f for pattern in ("*_GRiPHin.xlsx", "*_old_GRiPHin.xlsx") for f in glob.glob(pattern)]
            if len(griphin_files) < 2:
                raise ValueError(f"{CRED}Need at least two GRiPHin files for combination when using --griphin_list.{CEND}")
        else:
            # A file was provided, read it and get the paths from the file
            griphin_files = []
            with open(args.griphin_list, 'r') as file:
                for line in file:
                    line = line.strip()
                    if line:  # Skip empty lines
                        if os.path.isfile(line):  # Ensure the path exists
                            griphin_files.append(line)
                        else:
                            print(CRED + f"File '{line}' does not exist."+ CEND)
        print(CYELLOW + f"Combining '{griphin_files}' into one excel file."+ CEND)
        # Initialize the combination with the first file
        base_file = griphin_files.pop(0)
        #combine first two files
        combined_df_qc_final, combined_df_ar_final, combined_df_pf_final, combined_df_hv_final, phoenix_final, shiga_final, centar_final, ordered_centar_df_final, centar_df_lens_final, centar_df_column_names_final = read_excels(base_file, griphin_files[0], args.samplesheet, False, args.parent_folder)
        #print_df(combined_df_qc_final, "G ------- Combined QC DataFrame -------", False)
        #df_has_other_dupes(combined_df_qc_final, 'WGS_ID', 'G Combined QC DataFrame')
        combined_dataframes_final = [ combined_df_qc_final, combined_df_ar_final, combined_df_pf_final, combined_df_hv_final, ordered_centar_df_final ]
        # Iterate over remaining files, progressively combining them with the base
        for count, next_file in enumerate(griphin_files[1:], start=1):
            # checking
            combined_df_qc_next, combined_df_ar_next, combined_df_pf_next, combined_df_hv_next, phoenix, shiga, centar, ordered_centar_df_next, centar_df_lens_next, centar_df_column_names_next = read_excel(next_file, phoenix_final, combined_df_qc_final, ordered_centar_df_final, args.samplesheet, centar_df_lens_final, centar_df_column_names_final, args.parent_folder)
            combined_dataframes_next = [ combined_df_qc_next, combined_df_ar_next, combined_df_pf_next, combined_df_hv_next, ordered_centar_df_next ]
            # Update flags
            phoenix_final = phoenix_final or phoenix
            shiga_final = shiga_final or shiga
            centar_final = centar_final or centar
            combined_dataframes_final = [ add_and_concatenate(df1, df2, shiga_final) for df1, df2 in zip(combined_dataframes_final, combined_dataframes_next) ]
            if centar_final == True:
                # Update centar_df_lens with the maximum of each corresponding number
                #centar_df_lens_final = [max(lens) for lens in zip(centar_df_lens_final, centar_df_lens_next)]
                # Update to get columns names
                centar_df_lens_final, centar_df_column_names_final = update_centar_column_names_lens(centar_df_column_names_final, centar_df_column_names_next)
            # Unpack the tuple into individual DataFrames
            combined_df_qc_final, combined_df_ar_final, combined_df_pf_final, combined_df_hv_final, ordered_centar_df_final = combined_dataframes_final
    # add centar_df to the almost_final_ar_df
    if centar_final == True:
        # Reset the indices of both DataFrames during concat so they are aligned and we don't get NaNs in the final row of the dataframe
        ordered_centar_df_final = sort_columns_to_primary_ungrouped(ordered_centar_df_final)
        #combined_df_qc = pd.concat([combined_df_qc, ordered_centar_df], axis=1)
        #print_df(combined_df_qc_final, "H ------- Combined QC DataFrame with Centar -------", False)
        #df_has_other_dupes(combined_df_qc_final, 'WGS_ID', 'H Combined QC DataFrame with Centar')
        combined_df_qc_final = pd.merge(combined_df_qc_final, ordered_centar_df_final, how="left", on = ['UNI', 'UNI'])
        #print_df(combined_df_qc_final, "I ------- Combined QC DataFrame with Centar -------", False)
        #df_has_other_dupes(combined_df_qc_final, 'WGS_ID', 'I Combined QC DataFrame with Centar --- Post merge')
        combined_df_qc_final = sort_columns_to_primary_ungrouped(combined_df_qc_final)
    
    # call function from griphin script to combine all dfs
    final_df, ar_max_col, columns_to_highlight, final_ar_df, final_pf_db, final_ar_db, final_hv_db = Combine_dfs(combined_df_qc_final, combined_df_ar_final, combined_df_pf_final, combined_df_hv_final, pd.DataFrame(), phoenix_final, args.scaffolds, True, args.bldb)


    #print(list(final_df.index))
    #print(final_df['WGS_ID'].tolist())
    #print(final_df['UNI'].tolist())
    #check if we need to add shiga pass information
    #get other information for excel writing
    combined_df_qc_final = combined_df_qc_final.drop('UNI', axis = 1)
    (qc_max_row, qc_max_col) = combined_df_qc_final.shape
    pf_max_col = combined_df_pf_final.shape[1] - 1 #remove one for the UNI column
    hv_max_col = combined_df_hv_final.shape[1] - 1 #remove one for the UNI column
    #write excel sheet
    final_df = final_df.drop('UNI', axis=1)
    final_df = sort_samples_df(final_df)
    #df_has_index_dupes(final_df, "Final DataFrame after dropping UNI")
    #df_has_other_dupes(final_df, 'UNI', 'Final_Dataframe after dropping UNI')
    #df_has_other_dupes(final_df, 'WGS_ID', 'Final Dataframe after dropping UNI')
    write_to_excel(args.set_coverage, output_file, final_df, qc_max_col, ar_max_col, pf_max_col, hv_max_col, columns_to_highlight, final_ar_df, final_pf_db, final_ar_db, final_hv_db, phoenix_final, shiga_final, centar_final, centar_df_lens_final)
    #write tsv from excel
    convert_excel_to_tsv(output_file)

if __name__ == '__main__':
    main()
